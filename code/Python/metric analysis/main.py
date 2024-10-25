import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import kendalltau


metric_filepath = "1.xlsx"
correlation_analysis_filepath = "Correlation Analysis.xlsx"
consistency_analysis_filepath = "Consistency Analysis.xlsx"
coefficient_of_variation_analysis_filepath = "Coefficient of Variation Analysis.xlsx"


wb = openpyxl.load_workbook(metric_filepath)
wb.active = 0
ws = wb.active
col_header = ws["1"][1:]

N = len(col_header)
kendall_mat = np.diag(np.ones(N)/2).astype(np.float32)

for i in range(N-1):
    for j in range(i+1, N):
        X = ws[col_header[i].column_letter][1:]
        X = [item.value for item in X]
        Y = ws[col_header[j].column_letter][1:]
        Y = [item.value for item in Y]
        kendall_mat[i, j] = kendalltau(X, Y).statistic
kendall_mat = kendall_mat+kendall_mat.T

new_wb = openpyxl.Workbook()
ws = new_wb.active
for i in range(N+1):
    for j in range(N+1):
        if i == 0:
            if j != 0:
                ws.cell(row=i+1, column=j+1, value=col_header[j-1].value)
        else:
            if j == 0:
                ws.cell(row=i+1, column=j+1, value=col_header[i-1].value)
            else:
                ws.cell(row=i+1, column=j+1, value=kendall_mat[i-1, j-1])
new_wb.save(correlation_analysis_filepath)

ws = wb.active
M = len(ws["A"])-1
ranks = np.zeros((M, N+2), dtype=np.float32)
for j in range(N):
    X = ws[col_header[j].column_letter][1:]
    X = [item.value for item in X]
    X = pd.DataFrame(X)
    ranks[:, j] = X.rank(ascending=False).values[:, 0]

for i in range(M):
    ranks[i, -2] = np.sum(M-ranks[i, 0:-2])

X = ranks[:, -2]
X = pd.DataFrame(X)
ranks[:, -1] = X.rank(ascending=False).values[:, 0]

kendall_mat = np.zeros(N, dtype=np.float32)
for j in range(N):
    kendall_mat[j] = kendalltau(ranks[:, j], ranks[:, -1]).statistic

new_wb = openpyxl.Workbook()
ws = new_wb.active
for i in range(2):
    for j in range(N):
        if i == 0:
            ws.cell(row=i+1, column=j+1, value=col_header[j].value)
        else:
            ws.cell(row=i+1, column=j+1, value=kendall_mat[j])
new_wb.save(consistency_analysis_filepath)

new_wb = openpyxl.Workbook()
ws = wb.active
new_ws = new_wb.active
for i in range(2):
    for j in range(N):
        if i == 0:
            new_ws.cell(row=i+1, column=j+1, value=col_header[j].value)
        else:
            X = ws[col_header[j].column_letter][1:]
            X = np.array([item.value for item in X])
            new_ws.cell(row=i+1, column=j+1, value=X.std()/X.mean())
new_wb.save(coefficient_of_variation_analysis_filepath)